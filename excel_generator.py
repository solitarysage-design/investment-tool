"""
週次投資判断シート Excel 生成モジュール

シート構成:
  Sheet1: 保有銘柄一覧  — 評価損益・配当利回り・PBR など現状確認
  Sheet2: 新規候補銘柄  — スクリーニング通過銘柄（配当利回り順）
  Sheet3: 比較分析      — 保有銘柄 vs 候補銘柄の PBR・配当利回り散布図用データ
"""

import io
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import ScatterChart, Reference, Series


# ---------------------------------------------------------------------------
# カラーパレット
# ---------------------------------------------------------------------------
C_HEADER_HOLD  = "1F4E79"   # 保有シートヘッダー：濃紺
C_HEADER_CAND  = "375623"   # 候補シートヘッダー：濃緑
C_HEADER_COMP  = "4A235A"   # 比較シートヘッダー：濃紫
C_GAIN         = "C6EFCE"   # 含み益セル背景（薄緑）
C_LOSS         = "FFC7CE"   # 含み損セル背景（薄赤）
C_GAIN_FONT    = "276221"
C_LOSS_FONT    = "9C0006"
C_ODD          = "F2F2F2"   # 奇数行背景（薄グレー）
C_EVEN         = "FFFFFF"   # 偶数行背景（白）
C_ACCENT       = "FFD966"   # アクセント（黄）
C_BORDER       = "BFBFBF"


def _thin_border():
    side = Side(style="thin", color=C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill(color: str):
    return PatternFill("solid", fgColor=color)


def _header_font():
    return Font(bold=True, color="FFFFFF", size=10)


def _cell_font(bold=False, size=10):
    return Font(bold=bold, size=size)


def _fill(color: str):
    return PatternFill("solid", fgColor=color)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _right():
    return Alignment(horizontal="right", vertical="center")


def _auto_width(ws, min_width=8, max_width=40):
    """列幅を内容に合わせて自動調整"""
    for col_cells in ws.columns:
        length = 0
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(length + 2, max_width))


def _write_header(ws, row: int, headers: list[str], bg_color: str):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = _header_fill(bg_color)
        cell.font = _header_font()
        cell.alignment = _center()
        cell.border = _thin_border()


def _write_data_row(ws, row: int, values: list, alt: bool = False):
    bg = C_ODD if alt else C_EVEN
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _fill(bg)
        cell.font = _cell_font()
        cell.border = _thin_border()
        if isinstance(val, (int, float)):
            cell.alignment = _right()
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# 公開関数
# ---------------------------------------------------------------------------

def create_investment_excel(
    holdings_df: pd.DataFrame,
    candidates_df: pd.DataFrame,
    output_path: "str | Path | io.BytesIO",
) -> "Path | io.BytesIO":
    """
    投資判断シート Excel を生成して保存する。

    Parameters
    ----------
    holdings_df  : 保有銘柄 DataFrame（pdf_parser + jquants 情報付加済み）
    candidates_df: スクリーニング通過銘柄 DataFrame
    output_path  : 出力先パス (.xlsx) または io.BytesIO（Streamlit用）

    Returns
    -------
    Path または BytesIO
    """
    # BytesIO の場合はディレクトリ作成をスキップ
    if not isinstance(output_path, io.BytesIO):
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    _write_sheet1_holdings(wb, holdings_df)
    _write_sheet2_candidates(wb, candidates_df)
    _write_sheet3_comparison(wb, holdings_df, candidates_df)

    try:
        wb.save(output_path)
    except PermissionError:
        raise PermissionError(
            f"{Path(output_path).name} が開いています。Excelを閉じてから再実行してください。"
        )
    return output_path


# ---------------------------------------------------------------------------
# Sheet 1: 保有銘柄一覧
# ---------------------------------------------------------------------------

_HOLD_HEADERS = [
    "銘柄コード", "銘柄名", "口座", "保有株数",
    "平均取得単価", "現在値", "評価額(円)",
    "評価損益(円)", "評価損益率(%)",
    "配当利回り(%)", "PBR(倍)", "市場",
]

_HOLD_COL_MAP = {
    "銘柄コード":   "code",
    "銘柄名":       "name",
    "口座":         "account_type",
    "保有株数":     "quantity",
    "平均取得単価": "avg_cost",
    "現在値":       "current_price",
    "評価額(円)":   "assessed_value",
    "評価損益(円)": "unrealized_pl",
    "評価損益率(%)":"unrealized_pct",
    "配当利回り(%)":"div_yield",
    "PBR(倍)":      "pbr",
    "市場":         "market",
}

_HOLD_NUM_FMT = {
    "平均取得単価": "#,##0",
    "現在値":       "#,##0",
    "評価額(円)":   "#,##0",
    "評価損益(円)": '#,##0;[Red]-#,##0',
    "評価損益率(%)":"0.00%",
    "配当利回り(%)":"0.00%",
    "PBR(倍)":      "0.00",
}


def _write_sheet1_holdings(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("①保有銘柄一覧")
    ws.freeze_panes = "A2"

    # タイトル
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"保有銘柄一覧  （更新日: {datetime.now().strftime('%Y/%m/%d')}）"
    title_cell.font = Font(bold=True, size=12, color=C_HEADER_HOLD)
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 20

    # ヘッダー
    _write_header(ws, 2, _HOLD_HEADERS, C_HEADER_HOLD)

    if df.empty:
        ws.append(["データなし"])
        return

    # データ行
    for i, (_, row) in enumerate(df.iterrows()):
        r = i + 3
        values = []
        for h in _HOLD_HEADERS:
            col_key = _HOLD_COL_MAP.get(h)
            val = row.get(col_key) if col_key and col_key in row.index else None
            # パーセント列は小数に変換（pdfから%が入っている場合）
            if h in ("評価損益率(%)", "配当利回り(%)") and val is not None:
                val = _to_pct_decimal(val)
            values.append(val)
        _write_data_row(ws, r, values, alt=(i % 2 == 1))

        # 評価損益のセルに色を付ける（赤/緑）
        pl_col = _HOLD_HEADERS.index("評価損益(円)") + 1
        pl_cell = ws.cell(row=r, column=pl_col)
        pl_val = pl_cell.value
        if pl_val is not None:
            if pl_val >= 0:
                pl_cell.fill = _fill(C_GAIN)
                pl_cell.font = Font(color=C_GAIN_FONT, size=10)
            else:
                pl_cell.fill = _fill(C_LOSS)
                pl_cell.font = Font(color=C_LOSS_FONT, size=10)

    # 数値フォーマット適用
    _apply_number_format(ws, _HOLD_HEADERS, _HOLD_NUM_FMT, start_row=3)

    # 合計行
    last_data_row = 2 + len(df)
    summary_row = last_data_row + 2
    ws.cell(row=summary_row, column=1).value = "合計"
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    for h in ["評価額(円)", "評価損益(円)"]:
        col_idx = _HOLD_HEADERS.index(h) + 1
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.value = f"=SUM({col_letter}3:{col_letter}{last_data_row})"
        cell.font = Font(bold=True)
        cell.fill = _fill(C_ACCENT)
        cell.number_format = "#,##0"

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 2: 新規候補銘柄
# ---------------------------------------------------------------------------

_CAND_HEADERS = [
    "銘柄コード", "銘柄名", "市場", "業種",
    "現在値(円)", "PBR(倍)", "配当利回り(%)",
    "時価総額(億円)", "年間DPS(円)", "判定期間",
]

_CAND_COL_MAP = {
    "銘柄コード":     "code",
    "銘柄名":         "name",
    "市場":           "market",
    "業種":           "sector17",
    "現在値(円)":     "price",
    "PBR(倍)":        "pbr",
    "配当利回り(%)":  "div_yield",
    "時価総額(億円)": "market_cap",
    "年間DPS(円)":    "dps",
    "判定期間":       "period_type",
}

_CAND_NUM_FMT = {
    "現在値(円)":     "#,##0",
    "PBR(倍)":        "0.00",
    "配当利回り(%)":  "0.00%",
    "時価総額(億円)": "#,##0",
    "年間DPS(円)":    "#,##0.0",
}


def _write_sheet2_candidates(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("②新規候補銘柄")
    ws.freeze_panes = "A2"

    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"スクリーニング結果  （PBR≦{_pbr_label()}倍・配当利回り≧{_yield_label()}%・時価総額≧100億・減配なし）"
        f"  {datetime.now().strftime('%Y/%m/%d')} 現在"
    )
    title_cell.font = Font(bold=True, size=12, color=C_HEADER_CAND)
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 20

    _write_header(ws, 2, _CAND_HEADERS, C_HEADER_CAND)

    if df.empty:
        ws.cell(row=3, column=1).value = "スクリーニング条件を満たす銘柄がありませんでした"
        return

    # 配当利回り降順ソート
    sort_col = "div_yield" if "div_yield" in df.columns else df.columns[0]
    df_sorted = df.sort_values(sort_col, ascending=False).reset_index(drop=True)

    for i, (_, row) in enumerate(df_sorted.iterrows()):
        r = i + 3
        values = []
        for h in _CAND_HEADERS:
            col_key = _CAND_COL_MAP.get(h)
            val = row.get(col_key) if col_key and col_key in row.index else None
            if h == "時価総額(億円)" and val is not None:
                val = round(val / 1e8, 1)  # 円 → 億円
            if h == "配当利回り(%)" and val is not None:
                val = _to_pct_decimal(val)
            values.append(val)
        _write_data_row(ws, r, values, alt=(i % 2 == 1))

        # 配当利回り列を色分け（高いほど濃い緑）
        yield_col_idx = _CAND_HEADERS.index("配当利回り(%)") + 1
        yield_cell = ws.cell(row=r, column=yield_col_idx)
        if yield_cell.value and isinstance(yield_cell.value, float):
            pct = yield_cell.value * 100
            if pct >= 5.0:
                yield_cell.fill = _fill("00B050")
                yield_cell.font = Font(color="FFFFFF", size=10, bold=True)
            elif pct >= 4.0:
                yield_cell.fill = _fill("92D050")
            elif pct >= 3.0:
                yield_cell.fill = _fill("C6EFCE")

        # PBR列を色分け（低いほど濃い青）
        pbr_col_idx = _CAND_HEADERS.index("PBR(倍)") + 1
        pbr_cell = ws.cell(row=r, column=pbr_col_idx)
        if pbr_cell.value and isinstance(pbr_cell.value, float):
            pbr = pbr_cell.value
            if pbr <= 0.8:
                pbr_cell.fill = _fill("1F4E79")
                pbr_cell.font = Font(color="FFFFFF", size=10, bold=True)
            elif pbr <= 1.0:
                pbr_cell.fill = _fill("9DC3E6")
            elif pbr <= 1.3:
                pbr_cell.fill = _fill("DDEBF7")

    _apply_number_format(ws, _CAND_HEADERS, _CAND_NUM_FMT, start_row=3)
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 3: 比較分析
# ---------------------------------------------------------------------------

def _write_sheet3_comparison(
    wb: Workbook,
    holdings_df: pd.DataFrame,
    candidates_df: pd.DataFrame,
):
    ws = wb.create_sheet("③保有vs候補比較")
    ws.freeze_panes = "A2"

    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"保有銘柄 vs 新規候補  PBR・配当利回り比較  {datetime.now().strftime('%Y/%m/%d')}"
    title_cell.font = Font(bold=True, size=12, color=C_HEADER_COMP)
    title_cell.alignment = _center()

    comp_headers = [
        "区分", "銘柄コード", "銘柄名",
        "現在値(円)", "PBR(倍)", "配当利回り(%)", "時価総額(億円)", "業種"
    ]
    _write_header(ws, 2, comp_headers, C_HEADER_COMP)

    rows_written = 0

    def add_rows(df, label, bg_color):
        nonlocal rows_written
        if df.empty:
            return
        price_col = "current_price" if "current_price" in df.columns else "price"
        for _, row in df.iterrows():
            r = rows_written + 3
            div_yield = row.get("div_yield")
            if div_yield is not None:
                div_yield = _to_pct_decimal(div_yield)
            mktcap = row.get("market_cap")
            if mktcap is not None:
                mktcap = round(mktcap / 1e8, 1)
            values = [
                label,
                row.get("code"),
                row.get("name"),
                row.get(price_col),
                row.get("pbr"),
                div_yield,
                mktcap,
                row.get("sector17"),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = _fill(bg_color)
                cell.font = _cell_font()
                cell.border = _thin_border()
                if isinstance(val, (int, float)):
                    cell.alignment = _right()
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            rows_written += 1

    add_rows(holdings_df, "保有中", "DDEBF7")  # 薄青
    add_rows(candidates_df.head(30), "候補", "E2EFDA")  # 薄緑

    # 数値フォーマット
    _apply_number_format(ws, comp_headers, {
        "現在値(円)":     "#,##0",
        "PBR(倍)":        "0.00",
        "配当利回り(%)":  "0.00%",
        "時価総額(億円)": "#,##0",
    }, start_row=3)

    # 散布図: PBR vs 配当利回り
    _add_scatter_chart(ws, rows_written, len(holdings_df))

    _auto_width(ws)


def _add_scatter_chart(ws, total_rows: int, holdings_count: int):
    """PBR vs 配当利回りの散布図を追加"""
    if total_rows < 2:
        return

    chart = ScatterChart()
    chart.title = "PBR vs 配当利回り（◆保有  ●候補）"
    chart.style = 10
    chart.x_axis.title = "PBR (倍)"
    chart.y_axis.title = "配当利回り (%)"
    chart.x_axis.numFmt = "0.00"
    chart.y_axis.numFmt = "0.00"
    chart.width = 18
    chart.height = 14

    data_start_row = 3
    pbr_col = 5   # E列
    yield_col = 6  # F列

    # 保有銘柄シリーズ
    if holdings_count > 0:
        xvalues_hold = Reference(ws, min_col=pbr_col,
                                  min_row=data_start_row,
                                  max_row=data_start_row + holdings_count - 1)
        yvalues_hold = Reference(ws, min_col=yield_col,
                                  min_row=data_start_row,
                                  max_row=data_start_row + holdings_count - 1)
        series_hold = Series(yvalues_hold, xvalues_hold, title="保有銘柄")
        series_hold.marker.symbol = "diamond"
        series_hold.marker.size = 8
        series_hold.graphicalProperties.line.noFill = True
        chart.series.append(series_hold)

    # 候補銘柄シリーズ
    cand_start = data_start_row + holdings_count
    cand_end = data_start_row + total_rows - 1
    if cand_end >= cand_start:
        xvalues_cand = Reference(ws, min_col=pbr_col,
                                  min_row=cand_start, max_row=cand_end)
        yvalues_cand = Reference(ws, min_col=yield_col,
                                  min_row=cand_start, max_row=cand_end)
        series_cand = Series(yvalues_cand, xvalues_cand, title="新規候補")
        series_cand.marker.symbol = "circle"
        series_cand.marker.size = 6
        series_cand.graphicalProperties.line.noFill = True
        chart.series.append(series_cand)

    ws.add_chart(chart, "J2")


# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------

def _apply_number_format(ws, headers: list[str], fmt_map: dict, start_row: int):
    """数値列に書式を適用"""
    for h, fmt in fmt_map.items():
        if h not in headers:
            continue
        col_idx = headers.index(h) + 1
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter][start_row - 1:]:
            if cell.value is not None:
                cell.number_format = fmt


def _to_pct_decimal(val) -> float | None:
    """パーセント値を小数に変換（例: 3.5 → 0.035）
    PDFから読んだ値（3.5）と計算済み値（0.035）の両方に対応"""
    if val is None:
        return None
    try:
        f = float(val)
        # 1より大きければパーセント表記 → 小数に変換
        return f / 100 if f > 1 else f
    except (TypeError, ValueError):
        return None


def _pbr_label() -> str:
    try:
        import config
        return str(config.SCREEN_PBR_MAX)
    except Exception:
        return "1.5"


def _yield_label() -> str:
    try:
        import config
        return str(config.SCREEN_YIELD_MIN)
    except Exception:
        return "2.5"
